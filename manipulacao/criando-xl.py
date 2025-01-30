from openpyxl import Workbook

from openpyxl import load_workbook #lê o arquivo excel

# Criando um documento excel

wb = Workbook()

ws = wb.active
ws.title = 'Planilha de exemplo'

ws.append(['nome', 'idade', 'cidade'])
ws.append(['João', 30, 'São Paulo'])
ws.append(['Maria', 25, 'Rio de Janeiro'])

wb.save('exemplo.xlsx')


#ler o documento excel existente

wb = load_workbook('exemplo.xlsx')
ws = wb.active

for row in ws.iter_rows(values_only =True):
    print(row)

# Como editar uma planilha já existente

wb = load_workbook('exemplo.xlsx')
ws = wb.active
ws['A2'] = 'José' #aqui eu to alterando o valor da célula A2 dentro da planilha

wb.save('exemplo_editado.xlsx')
