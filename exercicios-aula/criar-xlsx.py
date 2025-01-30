'''
6. Criar uma planilha com dados fictícios
● Crie um programa que:
○ Gere um arquivo Excel chamado dados.xlsx.
○ Crie uma planilha chamada "Vendas".
○ Adicione cabeçalhos nas colunas (ex.: "Produto", "Quantidade", "Preço").
○ Adicione 5 linhas de dados fictícios.
○ Salve o arquivo.


'''

from openpyxl import Workbook
from openpyxl import load_workbook #lê o arquivo xlsx no terminal de python

while True: 
    print('Deseja adicionar os dados a planilha de vendas?\n1 - Sim\n2 - Não')
    decisao = int(input('- '))
    if decisao == 1:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Vendas'
        ws.append(['Produto', 'Quantidade', 'Preço'])
        for i in range(5):
            print(f'Insira os dados para a linha {i+1}: ')
            produto = input('nome do produto - ')
            quantidade = int(input('quantidade do produto - '))
            preco = float(input('preço do produto - '))
            ws.append([produto, quantidade, preco])
            wb.save('vendas.xlsx')
    elif decisao == 2:
        print('Ok, continuando o sistema\n')
        break
    else:
        print('Opção inválida...')

# 7 - abra o arquivo xlsx, leia o todo e imprima seus dados no terminal

wb = load_workbook('vendas.xlsx')
ws = wb.active
for row in ws.iter_rows(values_only = True):
    print('Dados contidos na planilha:\n')
    print(row)

# 8 - Filtrar dados dentro da panilha, leia a planilha e os dados de vendas, filtre as linhas no qual o valor do produto seja mais que R$ 100,00 - Salve essas linhas (ou essa linha n sei) em uma nova planilha chamada de vendas_filtradas.xlsx

while True:
    print('Deseja filtrar algum tipo de venda dentro da planilha?\n1 - Sim\n2 - Não')
    decisao = int(input('- '))
    
    if decisao == 1:
        wb = load_workbook('vendas.xlsx')
        ws = wb.active

        wb_new = load_workbook('vendas.xlsx')
        ws_new = wb_new.create_sheet('vendas_filtradas')
        for coluna in range(1, ws.max_column + 1):
            ws_new.cell(row = 1, column = coluna, value= ws.cell(row = 1, column= coluna).value) #o que é isso meu pai amado??? REVISAR URGENTE
        linha_new = 2
        for row in range(2, ws.max_column + 1): #o que é isso senhor Deus
            produto = ws.cell(row = row, column=1).value
            quantidade = ws.cell(row=row, column=2).value
            preco = ws.cell(row=row, column=3).value

            if preco > 100:
                for i in range(1, ws.max_column + 1):
                    ws_new.cell( row= linha_new, column= i, value = ws.cell(row = row, column=i).value)
                linha_new_denovo += 1

                wb_new.save('vendas_filtradas.xlsx')
    elif decisao == 2:
        print('Ok, saindo do sistema e continiuando o código...')
        break
    
    else:
        print('Opção inválida')
